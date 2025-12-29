import os
from typing import List
import logging
from docx import Document
import openpyxl
import re

# 尝试导入PyPDF2
try:
    import PyPDF2
except ImportError:
    try:
        import pypdf as PyPDF2  # 新版本可能使用pypdf
    except ImportError:
        PyPDF2 = None

logger = logging.getLogger(__name__)


class TextSplitter:
    """简单的文本分割器"""
    def __init__(self, chunk_size=1000, chunk_overlap=200):
        self.chunk_size = chunk_size
        self.chunk_overlap = chunk_overlap
    
    def split_text(self, text: str) -> List[str]:
        """将文本分割成块"""
        if not text:
            return []
        
        # 按段落和句子分割
        paragraphs = re.split(r'\n\s*\n', text)
        chunks = []
        current_chunk = ""
        
        for para in paragraphs:
            para = para.strip()
            if not para:
                continue
            
            # 如果当前块加上新段落超过大小，保存当前块
            if current_chunk and len(current_chunk) + len(para) > self.chunk_size:
                chunks.append(current_chunk)
                # 保留重叠部分
                if self.chunk_overlap > 0 and len(current_chunk) > self.chunk_overlap:
                    current_chunk = current_chunk[-self.chunk_overlap:] + "\n\n" + para
                else:
                    current_chunk = para
            else:
                if current_chunk:
                    current_chunk += "\n\n" + para
                else:
                    current_chunk = para
            
            # 如果单个段落就超过大小，强制分割
            if len(current_chunk) > self.chunk_size * 1.5:
                # 按句子分割
                sentences = re.split(r'[.!?。！？]\s+', current_chunk)
                temp_chunk = ""
                for sentence in sentences:
                    if len(temp_chunk) + len(sentence) > self.chunk_size:
                        if temp_chunk:
                            chunks.append(temp_chunk)
                        temp_chunk = sentence
                    else:
                        temp_chunk += " " + sentence if temp_chunk else sentence
                current_chunk = temp_chunk
        
        # 添加最后一个块
        if current_chunk:
            chunks.append(current_chunk)
        
        return chunks if chunks else [text]


class FileProcessor:
    def __init__(self):
        self.text_splitter = TextSplitter(
            chunk_size=1000,
            chunk_overlap=200
        )
    
    async def process_file(self, file_path: str, filename: str) -> List[str]:
        """
        处理文件并提取文本，然后分割成块
        """
        # 根据文件类型选择处理方法
        file_ext = os.path.splitext(filename)[1].lower()
        
        if file_ext == '.pdf':
            text = await self._extract_pdf(file_path)
        elif file_ext in ['.doc', '.docx']:
            text = await self._extract_docx(file_path)
        elif file_ext in ['.xlsx', '.xls']:
            text = await self._extract_excel(file_path)
        elif file_ext == '.txt':
            text = await self._extract_txt(file_path)
        else:
            raise ValueError(f"不支持的文件类型: {file_ext}")
        
        if not text:
            return []
        
        # 分割文本成块
        chunks = self.text_splitter.split_text(text)
        return chunks
    
    async def _extract_pdf(self, file_path: str) -> str:
        """提取PDF文本"""
        if PyPDF2 is None:
            raise Exception("PDF处理库未安装。请运行: pip install pypdf2")
        
        text = ""
        try:
            logger.info(f"开始提取PDF: {file_path}")
            
            # 检查文件是否存在
            if not os.path.exists(file_path):
                raise Exception(f"文件不存在: {file_path}")
            
            with open(file_path, 'rb') as file:
                try:
                    pdf_reader = PyPDF2.PdfReader(file)
                    total_pages = len(pdf_reader.pages)
                    logger.info(f"PDF总页数: {total_pages}")
                    
                    if total_pages == 0:
                        raise Exception("PDF文件没有页面")
                    
                    for i, page in enumerate(pdf_reader.pages):
                        try:
                            page_text = page.extract_text()
                            if page_text:
                                text += page_text + "\n"
                            logger.debug(f"提取第 {i+1}/{total_pages} 页，文本长度: {len(page_text) if page_text else 0}")
                        except Exception as e:
                            logger.warning(f"提取第 {i+1} 页时出错: {str(e)}")
                            continue
                except Exception as e:
                    # 尝试检查是否是加密PDF
                    if "encrypted" in str(e).lower() or "password" in str(e).lower():
                        raise Exception("PDF文件已加密，无法提取文本。请先解密PDF文件。")
                    raise
            
            logger.info(f"PDF提取完成，总文本长度: {len(text)}")
            if not text.strip():
                raise Exception("PDF文件中没有可提取的文本内容。可能是扫描版PDF（图片格式），需要使用OCR工具。")
        except Exception as e:
            logger.error(f"PDF提取失败: {str(e)}", exc_info=True)
            raise Exception(f"PDF提取失败: {str(e)}")
        return text
    
    async def _extract_docx(self, file_path: str) -> str:
        """提取Word文档文本"""
        text = ""
        try:
            doc = Document(file_path)
            for paragraph in doc.paragraphs:
                text += paragraph.text + "\n"
        except Exception as e:
            raise Exception(f"Word文档提取失败: {str(e)}")
        return text
    
    async def _extract_excel(self, file_path: str) -> str:
        """提取Excel文本"""
        text = ""
        try:
            workbook = openpyxl.load_workbook(file_path)
            for sheet_name in workbook.sheetnames:
                sheet = workbook[sheet_name]
                for row in sheet.iter_rows(values_only=True):
                    row_text = " ".join([str(cell) if cell else "" for cell in row])
                    if row_text.strip():
                        text += row_text + "\n"
        except Exception as e:
            raise Exception(f"Excel提取失败: {str(e)}")
        return text
    
    async def _extract_txt(self, file_path: str) -> str:
        """提取TXT文本"""
        try:
            with open(file_path, 'r', encoding='utf-8') as file:
                text = file.read()
        except UnicodeDecodeError:
            # 尝试其他编码
            with open(file_path, 'r', encoding='gbk') as file:
                text = file.read()
        except Exception as e:
            raise Exception(f"TXT文件读取失败: {str(e)}")
        return text

